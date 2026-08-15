from __future__ import annotations

from collections.abc import Iterator, Sequence
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import Order, ParsedWorkbook


class WorkbookValidationError(ValueError):
    """Raised when an export is unreadable or incomplete."""


FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "order_no": ("订单号", "订单单号"),
    "organization": ("所属组织", "执行组织"),
    "carrier": ("承运商名称",),
    "departed_at": ("离厂时间(承运商提货时间)", "离厂时间"),
    "wms_posted_at": ("WMS过账时间",),
    "expected_arrival_at": ("预计到达时间",),
    "transport_status": ("状态",),
    "contract_status": ("合同状态",),
    "box_count": ("总箱数",),
    "actual_arrival_at": ("实际到达时间",),
    "signed_at": ("签收时间",),
    "is_delayed": ("是否延迟", "是否延误"),
    "delay_reason": ("延迟原因", "延误原因"),
    "carrier_sla_hours": ("承运商时效", "配送时效"),
    "electronic_signed_at": ("电子签签署时间",),
    "detail_count": ("明细单总数", "订单数量"),
}

REQUIRED_FIELDS = {
    "order_no",
    "departed_at",
    "wms_posted_at",
    "transport_status",
    "contract_status",
    "box_count",
    "is_delayed",
    "delay_reason",
}


def read_orders(path: str | Path, *, expected_ui_total: int | None = None) -> ParsedWorkbook:
    source = Path(path).resolve()
    if not source.exists():
        raise WorkbookValidationError(f"Excel 文件不存在: {source}")
    if source.suffix.lower() not in {".xls", ".xlsx"}:
        raise WorkbookValidationError(f"仅支持 .xls/.xlsx: {source.name}")

    sheet_name, headers, rows = _read_rows(source)
    positions = _resolve_positions(headers)
    orders: list[Order] = []
    seen: set[str] = set()

    for row_number, row in enumerate(rows, start=2):
        if not any(not _is_empty(value) for value in row):
            continue
        order = _to_order(row, row_number, positions)
        if order.order_no in seen:
            raise WorkbookValidationError(
                f"订单号重复: {order.order_no}（第 {row_number} 行）"
            )
        seen.add(order.order_no)
        orders.append(order)

    if not orders:
        raise WorkbookValidationError("Excel 没有订单数据")
    if expected_ui_total is not None and len(orders) != expected_ui_total:
        raise WorkbookValidationError(
            f"页面显示 {expected_ui_total} 条，但 Excel 有 {len(orders)} 个唯一订单"
        )

    return ParsedWorkbook(source, sheet_name, tuple(headers), tuple(orders))


def _read_rows(path: Path) -> tuple[str, list[str], Iterator[Sequence[Any]]]:
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            header_values = next(iterator)
        except StopIteration as exc:
            raise WorkbookValidationError("Excel 是空文件") from exc
        headers = [_header(value) for value in header_values]
        rows = [tuple(row) for row in iterator]
        workbook.close()
        return sheet.title, headers, iter(rows)

    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise WorkbookValidationError("读取 .xls 需要安装 xlrd") from exc

    workbook = xlrd.open_workbook(path, on_demand=True)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise WorkbookValidationError("Excel 是空文件")
    headers = [_header(sheet.cell_value(0, column)) for column in range(sheet.ncols)]

    def iter_xls() -> Iterator[Sequence[Any]]:
        for row_index in range(1, sheet.nrows):
            yield tuple(
                _xls_value(sheet.cell(row_index, column), workbook.datemode)
                for column in range(sheet.ncols)
            )

    rows = list(iter_xls())
    workbook.release_resources()
    return sheet.name, headers, iter(rows)


def _xls_value(cell: Any, datemode: int) -> Any:
    import xlrd

    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, datemode)
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    return cell.value


def _resolve_positions(headers: Sequence[str]) -> dict[str, int | None]:
    normalized = {_header(value): index for index, value in enumerate(headers) if value}
    positions: dict[str, int | None] = {}
    missing: list[str] = []
    for internal, aliases in FIELD_ALIASES.items():
        position = next((normalized[alias] for alias in aliases if alias in normalized), None)
        positions[internal] = position
        if internal in REQUIRED_FIELDS and position is None:
            missing.append("/".join(aliases))
    if missing:
        raise WorkbookValidationError("缺少必要表头: " + ", ".join(missing))
    return positions


def _to_order(row: Sequence[Any], row_number: int, positions: dict[str, int | None]) -> Order:
    def value(name: str) -> Any:
        position = positions[name]
        return row[position] if position is not None and position < len(row) else None

    order_no = _text(value("order_no"))
    if not order_no:
        raise WorkbookValidationError(f"第 {row_number} 行订单号为空")
    departed_at = _datetime(value("departed_at"), "离厂时间", row_number)
    return Order(
        order_no=order_no,
        organization=_text(value("organization")),
        carrier=_text(value("carrier")),
        departed_at=departed_at,
        wms_posted_at=_datetime(value("wms_posted_at"), "WMS过账时间", row_number),
        expected_arrival_at=_datetime(
            value("expected_arrival_at"), "预计到达时间", row_number
        ),
        transport_status=_text(value("transport_status")),
        contract_status=_text(value("contract_status")),
        box_count=_integer(value("box_count"), "总箱数", row_number),
        actual_arrival_at=_datetime(value("actual_arrival_at"), "实际到达时间", row_number),
        signed_at=_datetime(value("signed_at"), "签收时间", row_number),
        is_delayed=_boolean(value("is_delayed"), "是否延迟", row_number),
        delay_reason=_optional_text(value("delay_reason")),
        carrier_sla_hours=_float(value("carrier_sla_hours")),
        electronic_signed_at=_datetime(
            value("electronic_signed_at"), "电子签签署时间", row_number
        ),
        detail_count=_optional_integer(value("detail_count")),
        source_row=row_number,
    )


def _header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _text(value: Any) -> str:
    if _is_empty(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _optional_text(value: Any) -> str | None:
    text = _text(value)
    return text or None


def _datetime(
    value: Any, field_name: str, row_number: int, *, required: bool = False
) -> datetime | None:
    if _is_empty(value):
        if required:
            raise WorkbookValidationError(f"第 {row_number} 行 {field_name} 为空")
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = str(value).strip()
    for fmt in (
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S.%f",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise WorkbookValidationError(f"第 {row_number} 行 {field_name} 无法解析: {text}")


def _integer(value: Any, field_name: str, row_number: int) -> int:
    if _is_empty(value):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise WorkbookValidationError(
            f"第 {row_number} 行 {field_name} 不是数字: {value}"
        ) from exc


def _optional_integer(value: Any) -> int | None:
    return None if _is_empty(value) else int(float(value))


def _float(value: Any) -> float | None:
    return None if _is_empty(value) else float(value)


def _boolean(value: Any, field_name: str, row_number: int) -> bool:
    normalized = _text(value).lower()
    if normalized in {"是", "true", "1", "yes", "y"}:
        return True
    if normalized in {"否", "false", "0", "no", "n", ""}:
        return False
    raise WorkbookValidationError(f"第 {row_number} 行 {field_name} 值未知: {value}")


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())
